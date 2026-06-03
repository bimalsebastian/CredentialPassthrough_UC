"""Tests for XLSX read and write handlers."""

import sys
import io
from unittest.mock import MagicMock, patch

import pytest
import pandas as pd

from conftest import (make_download_response, SparkSession, DataFrame,
                      pyspark_types, MockField, MockSchema)

from direct_adls_reader import DirectADLSReader
from direct_adls_writer import DirectADLSWriter, WriteTransactionContext


def _make_reader(mock_adls_client, spark):
    return DirectADLSReader(mock_adls_client, spark)


def _make_writer(mock_adls_client, spark):
    writer = object.__new__(DirectADLSWriter)
    writer._DirectADLSWriter__adls_client = mock_adls_client
    writer.spark = spark
    writer._DirectADLSWriter__lock = __import__("threading").Lock()
    writer._DirectADLSWriter__max_files_per_write = 1000
    writer._DirectADLSWriter__max_partition_files = 100
    writer._DirectADLSWriter__max_file_size_mb = 500
    writer._options = {}
    return writer


def _create_xlsx_bytes(data, sheet_name="Sheet1"):
    """Create a real XLSX file in memory for testing reads."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if data:
        headers = list(data[0].keys())
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        for row_idx, record in enumerate(data, 2):
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=record[h])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _import_blocker(blocked_module):
    """Return a side_effect function that blocks a specific import."""
    _real_import = __import__

    def _blocker(name, *args, **kwargs):
        if name == blocked_module or name.startswith(blocked_module + "."):
            raise ImportError(f"Mocked: {blocked_module} not installed")
        return _real_import(name, *args, **kwargs)

    return _blocker


# ===========================================================================
# READ TESTS
# ===========================================================================

class TestReadXlsxHappyPath:
    def test_reads_xlsx_active_sheet(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        xlsx_bytes = _create_xlsx_bytes([
            {"Name": "Alice", "Age": 30},
            {"Name": "Bob", "Age": 25},
        ])
        mock_file_client.download_file.return_value = make_download_response(xlsx_bytes)

        reader = _make_reader(mock_adls_client, spark)
        with patch.object(reader, "_resolve_file_paths", return_value=["data/people.xlsx"]):
            result = reader.read_xlsx_files("container", "data/people.xlsx")

        pdf = result.toPandas()
        assert "Name" in pdf.columns
        assert "Age" in pdf.columns
        assert len(pdf) == 2
        assert pdf["Name"].iloc[0] == "Alice"

    def test_reads_named_sheet(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Default"
        ws1.cell(row=1, column=1, value="wrong")

        ws2 = wb.create_sheet("Target")
        ws2.cell(row=1, column=1, value="Col1")
        ws2.cell(row=2, column=1, value="correct")

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        mock_file_client.download_file.return_value = make_download_response(buf.getvalue())

        reader = _make_reader(mock_adls_client, spark)
        with patch.object(reader, "_resolve_file_paths", return_value=["multi.xlsx"]):
            result = reader.read_xlsx_files("container", "multi.xlsx",
                                            options={"sheet_name": "Target"})

        pdf = result.toPandas()
        assert "Col1" in pdf.columns
        assert pdf["Col1"].iloc[0] == "correct"

    def test_strips_whitespace_from_strings(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        xlsx_bytes = _create_xlsx_bytes([{"Name": "  padded  "}])
        mock_file_client.download_file.return_value = make_download_response(xlsx_bytes)

        reader = _make_reader(mock_adls_client, spark)
        with patch.object(reader, "_resolve_file_paths", return_value=["ws.xlsx"]):
            result = reader.read_xlsx_files("container", "ws.xlsx")

        pdf = result.toPandas()
        assert pdf["Name"].iloc[0] == "padded"


class TestReadXlsxMissingDependency:
    def test_import_error_message(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        reader = _make_reader(mock_adls_client, spark)

        with patch.object(reader, "_resolve_file_paths", return_value=["file.xlsx"]):
            with patch("builtins.__import__", side_effect=_import_blocker("openpyxl")):
                with pytest.raises(RuntimeError, match="openpyxl is required for XLSX support"):
                    reader.read_xlsx_files("container", "file.xlsx")


class TestReadXlsxEmptyInput:
    def test_empty_xlsx_raises(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        from openpyxl import Workbook
        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        mock_file_client.download_file.return_value = make_download_response(buf.getvalue())

        reader = _make_reader(mock_adls_client, spark)
        with patch.object(reader, "_resolve_file_paths", return_value=["empty.xlsx"]):
            with pytest.raises(RuntimeError, match="No valid XLSX data"):
                reader.read_xlsx_files("container", "empty.xlsx")


# ===========================================================================
# WRITE TESTS
# ===========================================================================

class TestWriteXlsxHappyPath:
    def test_writes_xlsx_with_bold_headers(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        writer = _make_writer(mock_adls_client, spark)
        pdf = pd.DataFrame({"Name": ["Alice"], "Score": [95]})
        schema = MockSchema([
            MockField("Name", pyspark_types.StringType()),
            MockField("Score", pyspark_types.IntegerType()),
        ])
        df = DataFrame(pdf, schema)

        with patch.object(WriteTransactionContext, "__enter__", return_value=MagicMock(committed=False, get_temp_path=lambda: "tmp_xlsx")):
            with patch.object(WriteTransactionContext, "__exit__", return_value=False):
                writer.write_xlsx_files(df, "container", "output.xlsx", mode="overwrite")

        mock_file_client.upload_data.assert_called()
        uploaded_bytes = mock_file_client.upload_data.call_args[0][0]

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(uploaded_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=2, column=1).value == "Alice"
        wb.close()

    def test_custom_sheet_name(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        writer = _make_writer(mock_adls_client, spark)
        pdf = pd.DataFrame({"X": [1]})
        schema = MockSchema([MockField("X", pyspark_types.IntegerType())])
        df = DataFrame(pdf, schema)

        with patch.object(WriteTransactionContext, "__enter__", return_value=MagicMock(committed=False, get_temp_path=lambda: "tmp")):
            with patch.object(WriteTransactionContext, "__exit__", return_value=False):
                writer.write_xlsx_files(df, "container", "out.xlsx", mode="overwrite",
                                        options={"sheet_name": "MyData"})

        uploaded_bytes = mock_file_client.upload_data.call_args[0][0]
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(uploaded_bytes))
        assert wb.active.title == "MyData"
        wb.close()


class TestWriteXlsxMissingDependency:
    def test_import_error_message(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        writer = _make_writer(mock_adls_client, spark)
        pdf = pd.DataFrame({"X": [1]})
        schema = MockSchema([MockField("X", pyspark_types.IntegerType())])
        df = DataFrame(pdf, schema)

        with patch("builtins.__import__", side_effect=_import_blocker("openpyxl")):
            with pytest.raises(RuntimeError, match="openpyxl is required for XLSX support"):
                writer.write_xlsx_files(df, "container", "out.xlsx", mode="overwrite")


class TestWriteXlsxEmptyInput:
    def test_empty_dataframe_writes_headers_only(self, mock_adls_client, mock_file_system_client, mock_file_client, spark):
        writer = _make_writer(mock_adls_client, spark)
        pdf = pd.DataFrame({"Name": pd.Series(dtype=str)})
        schema = MockSchema([MockField("Name", pyspark_types.StringType())])
        df = DataFrame(pdf, schema)

        with patch.object(WriteTransactionContext, "__enter__", return_value=MagicMock(committed=False, get_temp_path=lambda: "tmp")):
            with patch.object(WriteTransactionContext, "__exit__", return_value=False):
                writer.write_xlsx_files(df, "container", "empty.xlsx", mode="overwrite")

        mock_file_client.upload_data.assert_called()
        uploaded_bytes = mock_file_client.upload_data.call_args[0][0]

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(uploaded_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=2, column=1).value is None
        wb.close()
