"""
Direct ADLS File Reader with DataFrame Conversion

This module reads unstructured files directly from ADLS using Python SDK with user credentials,
then converts the data to Spark DataFrames without requiring Spark-level token injection.

Required dependency: PyYAML (available in standard Databricks Runtime)
    pip install pyyaml

Required dependency: openpyxl (NOT included in Databricks Runtime by default —
    must be installed on the cluster via init script or %pip install openpyxl)

Optional dependency: mutagen (only needed for audio_mode='metadata')
    %pip install mutagen
    Audio binary mode (default) requires no extra dependencies.
"""

import io
import json
import logging
from typing import Optional, Dict, Any, List, Union, Iterator
from urllib.parse import urlparse
from datetime import datetime
import glob
import os

try:
    from pyspark.sql import DataFrame, SparkSession
    from pyspark.sql.types import StructType, StructField, StringType, BinaryType, LongType, TimestampType, IntegerType
    from pyspark.sql.functions import col, lit
    import pandas as pd
except ImportError as e:
    raise ImportError(f"Required libraries not found: {e}")

try:
    from azure.storage.filedatalake import DataLakeServiceClient, FileSystemClient, DataLakeFileClient
    from azure.core.exceptions import ResourceNotFoundError, HttpResponseError
    import chardet  # For encoding detection
except ImportError as e:
    raise ImportError(f"Required Azure libraries not found: {e}")

logger = logging.getLogger(__name__)

DEFAULT_CHUNK_SIZE_BYTES = 4 * 1024 * 1024  # 4MB


class DirectADLSReader:
    """
    Reads unstructured files directly from ADLS using Python SDK, then converts to Spark DataFrames.
    This bypasses the need for Spark-level credential injection.
    """
    
    def __init__(self, adls_client: DataLakeServiceClient, spark_session: SparkSession):
        """
        Initialize DirectADLSReader.
        
        Args:
            adls_client: Authenticated ADLS client with user credentials
            spark_session: Active Spark session for DataFrame creation
        """
        self.adls_client = adls_client
        self.spark = spark_session
        self.max_files_per_read = 1000  # Safety limit
        self.max_file_size_mb = 100  # Safety limit for individual files

    def _download_with_chunks(self, file_client, chunk_size_bytes: int = DEFAULT_CHUNK_SIZE_BYTES) -> bytes:
        stream = file_client.download_file(max_concurrency=4)
        buffer = io.BytesIO()
        for chunk in stream.chunks():
            buffer.write(chunk)
        buffer.seek(0)
        return buffer.read()

    def read_text_files(self, container: str, blob_path: str, 
                       encoding: Optional[str] = None,
                       options: Optional[Dict] = None) -> DataFrame:
        """
        Read text files directly from ADLS and convert to Spark DataFrame.
        
        Args:
            container: ADLS container name
            blob_path: Path to file(s) - supports wildcards
            encoding: File encoding (auto-detected if None)
            options: Additional reading options
            
        Returns:
            Spark DataFrame with text content
        """
        try:
            files_data = []
            file_system_client = self.adls_client.get_file_system_client(container)
            
            # Get list of files matching the pattern
            file_paths = self._resolve_file_paths(file_system_client, blob_path)
            
            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]
            
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    
                    # Get file properties
                    properties = file_client.get_file_properties()
                    file_size_mb = properties.size / (1024 * 1024)
                    
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} ({file_size_mb:.1f} MB)")
                        continue
                    
                    # Download file content
                    content_bytes = self._download_with_chunks(file_client)

                    # Detect encoding if not specified
                    if encoding is None:
                        detected = chardet.detect(content_bytes)
                        file_encoding = detected.get('encoding', 'utf-8')
                    else:
                        file_encoding = encoding
                    
                    # Decode content
                    content_text = content_bytes.decode(file_encoding)
                    
                    # Create file record
                    file_record = {
                        'path': f"abfss://{container}@{self.adls_client.account_name}.dfs.core.windows.net/{file_path}",
                        'modificationTime': properties.last_modified,
                        'length': properties.size,
                        'content': content_text,
                        'encoding': file_encoding
                    }
                    files_data.append(file_record)
                    
                except Exception as e:
                    logger.warning(f"Failed to read file {file_path}: {str(e)}")
                    continue
            
            if not files_data:
                raise RuntimeError(f"No files could be read from {blob_path}")
            
            # Convert to Spark DataFrame
            return self._create_text_dataframe(files_data)
            
        except Exception as e:
            logger.error(f"Failed to read text files from {blob_path}: {str(e)}")
            raise RuntimeError(f"Text file reading failed: {str(e)}")
    
    def read_binary_files(self, container: str, blob_path: str,
                         options: Optional[Dict] = None) -> DataFrame:
        """
        Read binary files directly from ADLS and convert to Spark DataFrame.
        
        Args:
            container: ADLS container name  
            blob_path: Path to file(s) - supports wildcards
            options: Additional reading options
            
        Returns:
            Spark DataFrame with binary content (similar to Spark's binaryFile format)
        """
        try:
            files_data = []
            file_system_client = self.adls_client.get_file_system_client(container)
            
            # Get list of files matching the pattern
            file_paths = self._resolve_file_paths(file_system_client, blob_path)
            
            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]
            
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    
                    # Get file properties
                    properties = file_client.get_file_properties()
                    file_size_mb = properties.size / (1024 * 1024)
                    
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} ({file_size_mb:.1f} MB)")
                        continue
                    
                    # Download file content
                    content_bytes = self._download_with_chunks(file_client)

                    # Create file record (similar to Spark's binaryFile format)
                    file_record = {
                        'path': f"abfss://{container}@{self.adls_client.account_name}.dfs.core.windows.net/{file_path}",
                        'modificationTime': properties.last_modified,
                        'length': properties.size,
                        'content': content_bytes
                    }
                    files_data.append(file_record)
                    
                except Exception as e:
                    logger.warning(f"Failed to read file {file_path}: {str(e)}")
                    continue
            
            if not files_data:
                raise RuntimeError(f"No files could be read from {blob_path}")
            
            # Convert to Spark DataFrame
            return self._create_binary_dataframe(files_data)
            
        except Exception as e:
            logger.error(f"Failed to read binary files from {blob_path}: {str(e)}")
            raise RuntimeError(f"Binary file reading failed: {str(e)}")
    
    def read_json_files(self, container: str, blob_path: str,
                       options: Optional[Dict] = None) -> DataFrame:
        """
        Read JSON files directly from ADLS and convert to Spark DataFrame.
        
        Args:
            container: ADLS container name
            blob_path: Path to file(s) - supports wildcards  
            options: Additional reading options (multiLine, etc.)
            
        Returns:
            Spark DataFrame with JSON data
        """
        try:
            all_json_data = []
            file_system_client = self.adls_client.get_file_system_client(container)
            
            # Get list of files matching the pattern
            file_paths = self._resolve_file_paths(file_system_client, blob_path)
            
            multiline = options.get('multiLine', False) if options else False
            
            for file_path in file_paths[:self.max_files_per_read]:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    
                    # Get file properties and check size
                    properties = file_client.get_file_properties()
                    file_size_mb = properties.size / (1024 * 1024)
                    
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} ({file_size_mb:.1f} MB)")
                        continue
                    
                    # Download and parse JSON content
                    content_bytes = self._download_with_chunks(file_client)
                    content_text = content_bytes.decode('utf-8')
                    
                    if multiline:
                        # Single JSON object per file
                        json_obj = json.loads(content_text)
                        json_obj['_file_path'] = f"abfss://{container}@{self.adls_client.account_name}.dfs.core.windows.net/{file_path}"
                        all_json_data.append(json_obj)
                    else:
                        # JSON Lines format (one JSON per line)
                        for line_num, line in enumerate(content_text.strip().split('\n')):
                            if line.strip():
                                try:
                                    json_obj = json.loads(line)
                                    json_obj['_file_path'] = f"abfss://{container}@{self.adls_client.account_name}.dfs.core.windows.net/{file_path}"
                                    json_obj['_line_number'] = line_num + 1
                                    all_json_data.append(json_obj)
                                except json.JSONDecodeError as e:
                                    logger.warning(f"Invalid JSON on line {line_num + 1} in {file_path}: {str(e)}")
                                    continue
                    
                except Exception as e:
                    logger.warning(f"Failed to read JSON file {file_path}: {str(e)}")
                    continue
            
            if not all_json_data:
                raise RuntimeError(f"No valid JSON data found in {blob_path}")
            
            # Convert to Pandas DataFrame first, then to Spark
            pandas_df = pd.json_normalize(all_json_data)
            spark_df = self.spark.createDataFrame(pandas_df)
            
            return spark_df
            
        except Exception as e:
            logger.error(f"Failed to read JSON files from {blob_path}: {str(e)}")
            raise RuntimeError(f"JSON file reading failed: {str(e)}")
    
    def read_csv_files(self, container: str, blob_path: str,
                      options: Optional[Dict] = None) -> DataFrame:
        """
        Read CSV files directly from ADLS and convert to Spark DataFrame.
        
        Args:
            container: ADLS container name
            blob_path: Path to file(s) - supports wildcards
            options: CSV reading options (header, sep, etc.)
            
        Returns:
            Spark DataFrame with CSV data
        """
        try:
            all_csv_data = []
            file_system_client = self.adls_client.get_file_system_client(container)
            
            # Get list of files matching the pattern
            file_paths = self._resolve_file_paths(file_system_client, blob_path)
            
            # CSV options
            csv_options = options or {}
            header = csv_options.get('header', True)
            separator = csv_options.get('sep', ',')
            
            for file_path in file_paths[:self.max_files_per_read]:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    
                    # Check file size
                    properties = file_client.get_file_properties()
                    file_size_mb = properties.size / (1024 * 1024)
                    
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} ({file_size_mb:.1f} MB)")
                        continue
                    
                    # Download and parse CSV content
                    content_bytes = self._download_with_chunks(file_client)

                    # Read CSV using pandas
                    csv_data = pd.read_csv(
                        io.BytesIO(content_bytes),
                        header=0 if header else None,
                        sep=separator,
                        **{k: v for k, v in csv_options.items() if k not in ['header', 'sep']}
                    )
                    
                    # Add file path column
                    csv_data['_file_path'] = f"abfss://{container}@{self.adls_client.account_name}.dfs.core.windows.net/{file_path}"
                    
                    all_csv_data.append(csv_data)
                    
                except Exception as e:
                    logger.warning(f"Failed to read CSV file {file_path}: {str(e)}")
                    continue
            
            if not all_csv_data:
                raise RuntimeError(f"No valid CSV data found in {blob_path}")
            
            # Concatenate all CSV data
            combined_df = pd.concat(all_csv_data, ignore_index=True)
            
            # Convert to Spark DataFrame
            spark_df = self.spark.createDataFrame(combined_df)
            
            return spark_df
            
        except Exception as e:
            logger.error(f"Failed to read CSV files from {blob_path}: {str(e)}")
            raise RuntimeError(f"CSV file reading failed: {str(e)}")
    
    def read_parquet_files(self, container: str, blob_path: str,
                        options: Optional[Dict] = None) -> DataFrame:
        """Read parquet files via PyArrow — no Spark ABFS driver involved."""
        try:
            import pyarrow.parquet as pq
            import pyarrow as pa

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            tables = []
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    content_bytes = self._download_with_chunks(file_client)
                    tables.append(pq.read_table(pa.BufferReader(content_bytes)))
                except Exception as e:
                    logger.warning(f"Failed to read parquet file {file_path}: {e}")

            if not tables:
                raise RuntimeError(f"No valid parquet data found in {blob_path}")

            import pyarrow as pa
            combined = pa.concat_tables(tables)
            return self.spark.createDataFrame(combined.to_pandas())

        except Exception as e:
            logger.error(f"Failed to read parquet files from {blob_path}: {e}")
            raise RuntimeError(f"Parquet file reading failed: {e}")


    def read_orc_files(self, container: str, blob_path: str,
                    options: Optional[Dict] = None) -> DataFrame:
        """Read ORC files via PyArrow — no Spark ABFS driver involved."""
        try:
            import pyarrow.orc as orc
            import io

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            all_data = []
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    content_bytes = self._download_with_chunks(file_client)
                    table = orc.read_table(io.BytesIO(content_bytes))
                    all_data.append(table.to_pandas())
                except Exception as e:
                    logger.warning(f"Failed to read ORC file {file_path}: {e}")

            if not all_data:
                raise RuntimeError(f"No valid ORC data found in {blob_path}")

            import pandas as pd
            return self.spark.createDataFrame(pd.concat(all_data, ignore_index=True))

        except ImportError:
            raise RuntimeError("PyArrow required for ORC reading: pip install pyarrow")
        except Exception as e:
            logger.error(f"Failed to read ORC files from {blob_path}: {e}")
            raise RuntimeError(f"ORC file reading failed: {e}")


    def read_avro_files(self, container: str, blob_path: str,
                        options: Optional[Dict] = None) -> DataFrame:
        """Read Avro files via fastavro — no Spark ABFS driver involved."""
        try:
            import fastavro
            import io
            import pandas as pd

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            all_records = []
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    raw = self._download_with_chunks(file_client)
                    if not isinstance(raw, (bytes, bytearray)):
                        raw = bytes(raw)
                    if len(raw) == 0:
                        logger.warning(f"Avro file is empty, skipping: {file_path}")
                        continue
                    buf = io.BytesIO(raw)
                    buf.seek(0)
                    records = list(fastavro.reader(buf))
                    all_records.extend(records)
                except Exception as e:
                    logger.warning(f"Failed to read Avro file {file_path}: {e}")

            if not all_records:
                raise RuntimeError(f"No valid Avro data found in {blob_path}")

            return self.spark.createDataFrame(pd.DataFrame(all_records))

        except ImportError:
            raise RuntimeError("fastavro required for Avro reading: pip install fastavro")
        except Exception as e:
            logger.error(f"Failed to read Avro files from {blob_path}: {e}")
            raise RuntimeError(f"Avro file reading failed: {e}")
    
    def read_xml_files(self, container: str, blob_path: str,
                    options: Optional[Dict] = None) -> DataFrame:
        """
        Read XML files via xml.etree — no Spark ABFS driver involved.
        Requires 'rowTag' option to identify the repeating element, 
        matching Spark's spark-xml behaviour.
        
        Example: options={'rowTag': 'person'}
        """
        try:
            import xml.etree.ElementTree as ET
            import pandas as pd

            options = options or {}
            row_tag = options.get('rowTag') or options.get('rowtag')
            if not row_tag:
                raise ValueError(
                    "XML reading requires 'rowTag' option specifying the repeating element. "
                    "Example: .option('rowTag', 'person')"
                )

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            all_records = []
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    content_bytes = self._download_with_chunks(file_client)
                    root = ET.fromstring(content_bytes.decode('utf-8'))

                    # Find all elements matching rowTag anywhere in the tree
                    elements = root.findall(f'.//{row_tag}')
                    if not elements:
                        # Maybe root itself is the row tag
                        if root.tag == row_tag:
                            elements = [root]

                    for elem in elements:
                        record = {}
                        # Attributes become columns
                        record.update(elem.attrib)
                        # Child elements become columns
                        for child in elem:
                            # Handle nested elements by flattening one level
                            if len(child) == 0:
                                record[child.tag] = child.text
                            else:
                                record[child.tag] = ET.tostring(child, encoding='unicode')
                        all_records.append(record)

                except Exception as e:
                    logger.warning(f"Failed to read XML file {file_path}: {e}")

            if not all_records:
                raise RuntimeError(f"No valid XML records found in {blob_path} "
                                f"with rowTag='{row_tag}'")

            return self.spark.createDataFrame(pd.DataFrame(all_records))

        except Exception as e:
            logger.error(f"Failed to read XML files from {blob_path}: {e}")
            raise RuntimeError(f"XML file reading failed: {e}")


    def read_yaml_files(self, container: str, blob_path: str,
                        options: Optional[Dict] = None) -> DataFrame:
        """
        Read YAML files from ADLS and return a Spark DataFrame with inferred schema.

        Nested keys are flattened using dot notation up to a maximum depth of 3 levels
        (e.g. config.database.host). Keys nested beyond depth 3 are serialised as a JSON
        string in a single column.

        Uses yaml.safe_load exclusively — yaml.load is never called.

        Args:
            container: ADLS container name
            blob_path: Path to file(s) - supports wildcards
            options: Additional reading options

        Returns:
            Spark DataFrame with flattened YAML content
        """
        try:
            import yaml

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, "
                               f"limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]

            all_records = []
            max_depth = 3

            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    properties = file_client.get_file_properties()

                    file_size_mb = properties.size / (1024 * 1024)
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} "
                                       f"({file_size_mb:.1f} MB)")
                        continue

                    content_bytes = self._download_with_chunks(file_client)
                    yaml_data = yaml.safe_load(content_bytes.decode('utf-8'))

                    if yaml_data is None:
                        logger.warning(f"Empty YAML file, skipping: {file_path}")
                        continue

                    # Handle both single-document dicts and lists of dicts
                    if isinstance(yaml_data, dict):
                        yaml_data = [yaml_data]
                    elif not isinstance(yaml_data, list):
                        yaml_data = [{'value': yaml_data}]

                    for doc in yaml_data:
                        if not isinstance(doc, dict):
                            doc = {'value': doc}
                        record = self._flatten_yaml_dict(doc, max_depth=max_depth)
                        record['_file_path'] = (
                            f"abfss://{container}@"
                            f"{self.adls_client.account_name}"
                            f".dfs.core.windows.net/{file_path}"
                        )
                        all_records.append(record)

                except Exception as e:
                    logger.warning(f"Failed to read YAML file {file_path}: {e}")
                    continue

            if not all_records:
                raise RuntimeError(f"No valid YAML data found in {blob_path}")

            return self.spark.createDataFrame(pd.DataFrame(all_records))

        except ImportError:
            raise RuntimeError("PyYAML required for YAML reading: pip install pyyaml")
        except Exception as e:
            logger.error(f"Failed to read YAML files from {blob_path}: {e}")
            raise RuntimeError(f"YAML file reading failed: {e}")

    def _flatten_yaml_dict(self, data: dict, prefix: str = '',
                           current_depth: int = 0, max_depth: int = 3) -> dict:
        """
        Flatten a nested dict using dot notation up to max_depth levels.
        Keys nested beyond max_depth are serialised as a JSON string.
        """
        result = {}
        for key, value in data.items():
            flat_key = f"{prefix}{key}" if not prefix else f"{prefix}.{key}"

            if isinstance(value, dict) and current_depth < max_depth - 1:
                nested = self._flatten_yaml_dict(
                    value, prefix=flat_key, current_depth=current_depth + 1,
                    max_depth=max_depth
                )
                result.update(nested)
            elif isinstance(value, (dict, list)) and current_depth >= max_depth - 1:
                # Beyond max depth — serialise as JSON string
                result[flat_key] = json.dumps(value)
            else:
                result[flat_key] = value

        return result

    def read_xlsx_files(self, container: str, blob_path: str,
                        options: Optional[Dict] = None) -> DataFrame:
        """
        Read Excel (.xlsx) files from ADLS and return a Spark DataFrame.

        Downloads file bytes into a BytesIO buffer, opens with openpyxl in read-only
        mode, reads the active sheet (or a named sheet via the 'sheet_name' option),
        handles merged cells by forward-filling, and strips whitespace from strings.

        Args:
            container: ADLS container name
            blob_path: Path to file(s) - supports wildcards
            options: Reading options. Supported keys:
                     - sheet_name: name of the worksheet to read (default: active sheet)
                     - header: whether first row is a header (default: True)

        Returns:
            Spark DataFrame with the spreadsheet contents
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for XLSX support. "
                "Install it with: %pip install openpyxl"
            )

        try:
            opts = options or {}
            sheet_name = opts.get('sheet_name') or opts.get('sheetName')
            header = opts.get('header', True)

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, "
                               f"limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]

            all_dataframes = []

            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    properties = file_client.get_file_properties()

                    file_size_mb = properties.size / (1024 * 1024)
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} "
                                       f"({file_size_mb:.1f} MB)")
                        continue

                    content_bytes = self._download_with_chunks(file_client)
                    wb = openpyxl.load_workbook(
                        io.BytesIO(content_bytes), read_only=True, data_only=True
                    )

                    # Select the target worksheet
                    if sheet_name:
                        if sheet_name not in wb.sheetnames:
                            logger.warning(
                                f"Sheet '{sheet_name}' not found in {file_path}, "
                                f"available: {wb.sheetnames}. Skipping."
                            )
                            wb.close()
                            continue
                        ws = wb[sheet_name]
                    else:
                        ws = wb.active

                    # Read all rows into a list of lists
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(list(row))
                    wb.close()

                    if not rows:
                        logger.warning(f"Empty worksheet in {file_path}, skipping")
                        continue

                    # Handle merged cells by forward-filling None values
                    rows = self._forward_fill_merged_cells(rows)

                    # Build pandas DataFrame
                    if header and len(rows) > 1:
                        columns = [
                            str(c).strip() if c is not None else f"_col{i}"
                            for i, c in enumerate(rows[0])
                        ]
                        pdf = pd.DataFrame(rows[1:], columns=columns)
                    else:
                        pdf = pd.DataFrame(rows)

                    # Strip whitespace from string columns
                    for col in pdf.select_dtypes(include=['object']).columns:
                        pdf[col] = pdf[col].apply(
                            lambda v: v.strip() if isinstance(v, str) else v
                        )

                    pdf['_file_path'] = (
                        f"abfss://{container}@"
                        f"{self.adls_client.account_name}"
                        f".dfs.core.windows.net/{file_path}"
                    )
                    all_dataframes.append(pdf)

                except Exception as e:
                    logger.warning(f"Failed to read XLSX file {file_path}: {e}")
                    continue

            if not all_dataframes:
                raise RuntimeError(f"No valid XLSX data found in {blob_path}")

            combined = pd.concat(all_dataframes, ignore_index=True)
            return self.spark.createDataFrame(combined)

        except RuntimeError:
            raise
        except Exception as e:
            logger.error(f"Failed to read XLSX files from {blob_path}: {e}")
            raise RuntimeError(f"XLSX file reading failed: {e}")

    @staticmethod
    def _forward_fill_merged_cells(rows: List[list]) -> List[list]:
        """
        Forward-fill None values across columns to handle merged cell ranges.
        A None that follows a non-None value in the same row is filled with
        the preceding value — this matches openpyxl's read_only behaviour where
        merged cells report None for all but the top-left cell.
        """
        for row in rows:
            for i in range(1, len(row)):
                if row[i] is None and row[i - 1] is not None:
                    row[i] = row[i - 1]
        return rows

    SUPPORTED_AUDIO_EXTENSIONS = {'wav', 'mp3', 'flac', 'aac', 'ogg', 'm4a'}

    def read_audio_files(self, container: str, blob_path: str,
                         options: Optional[Dict] = None) -> DataFrame:
        """
        Read audio files from ADLS with two modes controlled by the 'audio_mode' option.

        Mode 1 — "binary" (default):
            Returns a DataFrame with columns: path (StringType), content (BinaryType),
            filename (StringType). No additional dependencies required.

        Mode 2 — "metadata":
            Uses mutagen to extract audio metadata. Returns a DataFrame with columns:
            filename, format, duration_seconds, sample_rate, channels, bitrate,
            file_size_bytes. Requires mutagen to be installed.

        Supported extensions: wav, mp3, flac, aac, ogg, m4a

        Args:
            container: ADLS container name
            blob_path: Path to file(s) - supports wildcards
            options: Reading options. Supported keys:
                     - audio_mode: 'binary' (default) or 'metadata'

        Returns:
            Spark DataFrame (schema depends on audio_mode)
        """
        opts = options or {}
        audio_mode = opts.get('audio_mode', 'binary').lower()

        if audio_mode == 'metadata':
            return self._read_audio_metadata(container, blob_path, opts)
        else:
            return self._read_audio_binary(container, blob_path, opts)

    def _read_audio_binary(self, container: str, blob_path: str,
                           options: Dict) -> DataFrame:
        """Read audio files as raw binary content — no extra dependencies."""
        try:
            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, "
                               f"limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]

            files_data = []
            for file_path in file_paths:
                try:
                    # Validate audio extension
                    filename = file_path.split('/')[-1]
                    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
                    if ext not in self.SUPPORTED_AUDIO_EXTENSIONS:
                        logger.warning(f"Skipping non-audio file: {file_path}")
                        continue

                    file_client = file_system_client.get_file_client(file_path)
                    properties = file_client.get_file_properties()

                    file_size_mb = properties.size / (1024 * 1024)
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} "
                                       f"({file_size_mb:.1f} MB)")
                        continue

                    content_bytes = self._download_with_chunks(file_client)

                    files_data.append({
                        'path': (f"abfss://{container}@"
                                 f"{self.adls_client.account_name}"
                                 f".dfs.core.windows.net/{file_path}"),
                        'content': bytearray(content_bytes),
                        'filename': filename,
                    })

                except Exception as e:
                    logger.warning(f"Failed to read audio file {file_path}: {e}")

            if not files_data:
                raise RuntimeError(f"No audio files could be read from {blob_path}")

            schema = StructType([
                StructField("path",     StringType(), False),
                StructField("content",  BinaryType(), True),
                StructField("filename", StringType(), False),
            ])

            return self.spark.createDataFrame(files_data, schema)

        except RuntimeError:
            raise
        except Exception as e:
            logger.error(f"Failed to read audio files (binary) from {blob_path}: {e}")
            raise RuntimeError(f"Audio file reading (binary) failed: {e}")

    def _read_audio_metadata(self, container: str, blob_path: str,
                             options: Dict) -> DataFrame:
        """Read audio files and extract metadata using mutagen."""
        try:
            import mutagen
        except ImportError:
            raise RuntimeError(
                "mutagen is required for audio metadata mode. "
                "Install with: %pip install mutagen. "
                "For raw file access use audio_mode='binary' (no extra dependencies)."
            )

        try:
            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, "
                               f"limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]

            records = []
            for file_path in file_paths:
                try:
                    filename = file_path.split('/')[-1]
                    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
                    if ext not in self.SUPPORTED_AUDIO_EXTENSIONS:
                        logger.warning(f"Skipping non-audio file: {file_path}")
                        continue

                    file_client = file_system_client.get_file_client(file_path)
                    properties = file_client.get_file_properties()

                    file_size_mb = properties.size / (1024 * 1024)
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} "
                                       f"({file_size_mb:.1f} MB)")
                        continue

                    content_bytes = self._download_with_chunks(file_client)

                    # mutagen.File can read from a file-like object
                    audio_info = mutagen.File(io.BytesIO(content_bytes))

                    duration = None
                    sample_rate = None
                    channels = None
                    bitrate = None
                    audio_format = ext.upper()

                    if audio_info is not None and audio_info.info is not None:
                        duration = getattr(audio_info.info, 'length', None)
                        sample_rate = getattr(audio_info.info, 'sample_rate', None)
                        channels = getattr(audio_info.info, 'channels', None)
                        bitrate = getattr(audio_info.info, 'bitrate', None)

                    records.append({
                        'filename': filename,
                        'format': audio_format,
                        'duration_seconds': float(duration) if duration else None,
                        'sample_rate': int(sample_rate) if sample_rate else None,
                        'channels': int(channels) if channels else None,
                        'bitrate': int(bitrate) if bitrate else None,
                        'file_size_bytes': properties.size,
                    })

                except Exception as e:
                    logger.warning(f"Failed to read audio metadata for {file_path}: {e}")

            if not records:
                raise RuntimeError(f"No audio metadata could be extracted from {blob_path}")

            return self.spark.createDataFrame(pd.DataFrame(records))

        except RuntimeError:
            raise
        except Exception as e:
            logger.error(f"Failed to read audio metadata from {blob_path}: {e}")
            raise RuntimeError(f"Audio metadata reading failed: {e}")

    def read_binary_files(self, container: str, blob_path: str,
                        options: Optional[Dict] = None) -> DataFrame:
        """
        Read binary files and return a DataFrame matching Spark's binaryFile schema:
            path, modificationTime, length, content (bytes)
        """
        try:
            from pyspark.sql.types import (StructType, StructField, StringType,
                                        BinaryType, LongType, TimestampType)

            file_system_client = self.adls_client.get_file_system_client(container)
            file_paths = self._resolve_file_paths(file_system_client, blob_path)

            if len(file_paths) > self.max_files_per_read:
                logger.warning(f"Found {len(file_paths)} files, "
                            f"limiting to {self.max_files_per_read}")
                file_paths = file_paths[:self.max_files_per_read]

            files_data = []
            for file_path in file_paths:
                try:
                    file_client = file_system_client.get_file_client(file_path)
                    properties = file_client.get_file_properties()

                    file_size_mb = properties.size / (1024 * 1024)
                    if file_size_mb > self.max_file_size_mb:
                        logger.warning(f"Skipping large file {file_path} "
                                    f"({file_size_mb:.1f} MB)")
                        continue

                    content_bytes = self._download_with_chunks(file_client)

                    files_data.append({
                        'path': (f"abfss://{container}@"
                                f"{self.adls_client.account_name}"
                                f".dfs.core.windows.net/{file_path}"),
                        'modificationTime': properties.last_modified,
                        'length': properties.size,
                        'content': bytearray(content_bytes)   # bytearray for Spark BinaryType
                    })

                except Exception as e:
                    logger.warning(f"Failed to read binary file {file_path}: {e}")

            if not files_data:
                raise RuntimeError(f"No binary files could be read from {blob_path}")

            schema = StructType([
                StructField("path",             StringType(),    False),
                StructField("modificationTime", TimestampType(), True),
                StructField("length",           LongType(),      True),
                StructField("content",          BinaryType(),    True),
            ])

            return self.spark.createDataFrame(files_data, schema)

        except Exception as e:
            logger.error(f"Failed to read binary files from {blob_path}: {e}")
            raise RuntimeError(f"Binary file reading failed: {e}")
    
    def _resolve_file_paths(self, file_system_client: FileSystemClient, blob_path: str) -> List[str]:
        try:
            if '*' in blob_path or '?' in blob_path:
                return self._resolve_wildcard_paths(file_system_client, blob_path)
            
            # If path has a file extension — treat as file directly, skip metadata check
            filename = blob_path.split('/')[-1]
            if '.' in filename:
                return [blob_path]
            
            # No extension — could be a directory, check via get_paths
            try:
                paths = list(file_system_client.get_paths(path=blob_path, max_results=1))
                if paths:
                    # It's a directory — list all files inside
                    return self._list_directory_files(file_system_client, blob_path)
                else:
                    # Empty directory or single file with no extension
                    return [blob_path]
            except Exception:
                return [blob_path]

        except Exception as e:
            logger.error(f"Failed to resolve file paths for {blob_path}: {str(e)}")
            raise RuntimeError(f"Path resolution failed: {str(e)}")
    
    def _resolve_wildcard_paths(self, file_system_client: FileSystemClient, pattern: str) -> List[str]:
        """
        Resolve wildcard patterns to actual file paths.
        
        Args:
            file_system_client: ADLS file system client
            pattern: Path pattern with wildcards
            
        Returns:
            List of matching file paths
        """
        try:
            # Extract directory part and filename pattern
            pattern_parts = pattern.split('/')
            base_path = '/'.join(pattern_parts[:-1]) if len(pattern_parts) > 1 else ""
            filename_pattern = pattern_parts[-1]
            
            # List files in the base directory
            paths = file_system_client.get_paths(path=base_path, recursive=False)
            
            matching_files = []
            for path in paths:
                if not path.is_directory:
                    filename = path.name.split('/')[-1]
                    
                    # Simple wildcard matching (could be enhanced with fnmatch)
                    if self._matches_pattern(filename, filename_pattern):
                        matching_files.append(path.name)
            
            return sorted(matching_files)
            
        except Exception as e:
            logger.error(f"Failed to resolve wildcard pattern {pattern}: {str(e)}")
            return []
    
    def _list_directory_files(self, file_system_client: FileSystemClient, directory: str) -> List[str]:
        """
        List all files in a directory.
        
        Args:
            file_system_client: ADLS file system client
            directory: Directory path
            
        Returns:
            List of file paths in the directory
        """
        try:
            paths = file_system_client.get_paths(path=directory, recursive=False)
            
            file_paths = []
            for path in paths:
                if not path.is_directory:
                    file_paths.append(path.name)
            
            return sorted(file_paths)
            
        except Exception as e:
            logger.error(f"Failed to list directory {directory}: {str(e)}")
            return []
    
    def _matches_pattern(self, filename: str, pattern: str) -> bool:
        """
        Simple wildcard pattern matching.
        
        Args:
            filename: File name to test
            pattern: Pattern with * and ? wildcards
            
        Returns:
            True if filename matches pattern
        """
        import fnmatch
        return fnmatch.fnmatch(filename, pattern)
    
    def _create_text_dataframe(self, files_data: List[Dict]) -> DataFrame:
        """
        Create Spark DataFrame for text files.
        
        Args:
            files_data: List of file data dictionaries
            
        Returns:
            Spark DataFrame with text content
        """
        # Define schema for text files
        schema = StructType([
            StructField("path", StringType(), False),
            StructField("modificationTime", TimestampType(), False),
            StructField("length", LongType(), False),
            StructField("content", StringType(), True),
            StructField("encoding", StringType(), True)
        ])
        
        # Convert to Spark DataFrame
        return self.spark.createDataFrame(files_data, schema)
    
    def _create_binary_dataframe(self, files_data: List[Dict]) -> DataFrame:
        """
        Create Spark DataFrame for binary files (similar to binaryFile format).
        
        Args:
            files_data: List of file data dictionaries
            
        Returns:
            Spark DataFrame with binary content
        """
        # Define schema for binary files (matches Spark's binaryFile format)
        schema = StructType([
            StructField("path", StringType(), False),
            StructField("modificationTime", TimestampType(), False),
            StructField("length", LongType(), False),
            StructField("content", BinaryType(), True)
        ])
        
        # Convert to Spark DataFrame
        return self.spark.createDataFrame(files_data, schema)


# # Example usage and testing
# if __name__ == "__main__":
#     # Configure logging
#     logging.basicConfig(level=logging.INFO)
    
#     print("DirectADLSReader class created successfully")
#     print("This class reads files directly from ADLS and converts to Spark DataFrames")
#     print("\nSupported formats:")
#     print("- Text files: read_text_files()")
#     print("- Binary files: read_binary_files()")  
#     print("- JSON files: read_json_files()")
#     print("- CSV files: read_csv_files()")
#     print("\nExample usage:")
#     print("reader = DirectADLSReader(adls_client, spark)")
#     print("df = reader.read_text_files('container', 'raw/logs/*.log')")
#     print("df = reader.read_binary_files('container', 'images/*.jpg')")
#     print("df = reader.read_json_files('container', 'data/*.json', {'multiLine': True})")
